# Linux命令输出到Excel
# 导入相关库
import subprocess, datetime, os
from openpyxl import Workbook, load_workbook

# 修改工作路径
path = r"/home/bob/code/python"
os.chdir(path)

# 数据文件
file = 'test.xlsx'

# 执行命令
command1 = ['df', '-h']
stdout1 = subprocess.Popen(command1, stdout=subprocess.PIPE)

# grep上面的返回结果
command2 = ['grep', 'nvme0n']
stdout2 = subprocess.check_output(command2, stdin=stdout1.stdout)

# 转为数组
output = stdout2.split()

# 现在时间
time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

# 添加时间
row = [time, *output]
# 分解output数组，与time字符串组合后形成新数组
# row = [time] + output # 作用同上，速度比上面慢

# 显示将要写入的内容
print(row)

# 打开或者创建Excel空文档
if os.path.isfile(file): # 判断是否存在“file”
    workbook = load_workbook(file) # 如果是，打开这个“file”文件
else:
    workbook = Workbook() # 如果否，新建Excel

# 打开上次活动的工作簿
sheet = workbook.active

# 写入数据
sheet.append(row)

# 保存文件
workbook.save(file)