from openpyxl import load_workbook

wb = load_workbook('data.xlsx')
sheet = wb.active

# 读取数据
for row in sheet.rows:
    for cell in row:
        print(cell.value)

# 写入数据  
sheet.cell(row=1, column=1, value='Hello')
wb.save('output.xlsx')